"""Генератор акта проработки из шаблона Excel."""
from __future__ import annotations

import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from loguru import logger


# Путь к шаблону акта
TEMPLATE_PATH = Path(__file__).parent.parent.parent / "files" / "ШАБЛОН АКТ ПРОРАБОТКИ.xlsx"


@dataclass
class ActData:
    """Данные для заполнения акта проработки."""
    
    request_id: str  # ID заявки (REQ-XXXXX)
    date: str  # Дата проработки (ДД.ММ.ГГГГ)
    product_name: str  # Наименование товара поставщика
    supplier_name: str  # Поставщик (фирма, бренд, страна)
    iiko_product_name: str  # Наименование продукта из iiko
    
    # Опциональные поля
    production_date: Optional[str] = None  # Дата изготовления
    expiry_date: Optional[str] = None  # Срок годности


def generate_act(data: ActData, output_dir: Optional[Path] = None) -> Path:
    """
    Сгенерировать акт проработки из шаблона.
    
    Args:
        data: Данные для заполнения акта
        output_dir: Директория для сохранения (по умолчанию temp)
        
    Returns:
        Путь к сгенерированному файлу
    """
    logger.info(f"generate_act: request_id={data.request_id}, product={data.product_name[:30]}...")
    
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Шаблон акта не найден: {TEMPLATE_PATH}")
    
    # Создаём директорию для вывода
    if output_dir is None:
        output_dir = Path(tempfile.mkdtemp())
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Имя выходного файла
    safe_name = data.product_name.replace("/", "_").replace("\\", "_")[:50]
    output_filename = f"АКТ_ПРОРАБОТКИ_{data.request_id}_{safe_name}.xlsx"
    output_path = output_dir / output_filename
    
    # Копируем шаблон
    shutil.copy(TEMPLATE_PATH, output_path)
    logger.debug(f"Шаблон скопирован в {output_path}")
    
    # Открываем и заполняем
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active  # Лист "Акт"
    
    # Заменяем плейсхолдеры
    replacements = {
        "{{id_item}}": data.request_id,
        "{{date}}": data.date,
        "{{name_of_goods}}": data.product_name,
        "{{partner}}": data.supplier_name,
    }
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for placeholder, value in replacements.items():
                    if placeholder in cell.value:
                        cell.value = cell.value.replace(placeholder, value)
                        logger.debug(f"Заменён {placeholder} в ячейке {cell.coordinate}")
    
    # Заполняем наименование из iiko (строка 18-19)
    # Ищем ячейку после "Наименование полуфабриката (продукта) из iiko"
    for row_idx in range(18, 21):
        cell = ws.cell(row=row_idx, column=3)  # Колонка C
        if cell.value is None or cell.value == "":
            cell.value = data.iiko_product_name
            logger.debug(f"Записано наименование iiko в C{row_idx}")
            break
    
    # Сохраняем
    wb.save(output_path)
    wb.close()
    
    logger.info(f"Акт сгенерирован: {output_path}")
    return output_path


def generate_act_for_request(
    request_id: str,
    product_name: str,
    supplier_name: str,
    iiko_product_name: str,
    output_dir: Optional[Path] = None,
) -> Path:
    """
    Упрощённая функция генерации акта.
    
    Args:
        request_id: ID заявки
        product_name: Название товара поставщика
        supplier_name: Название поставщика
        iiko_product_name: Название продукта из iiko
        output_dir: Директория для сохранения
        
    Returns:
        Путь к файлу акта
    """
    data = ActData(
        request_id=request_id,
        date=datetime.now().strftime("%d.%m.%Y"),
        product_name=product_name,
        supplier_name=supplier_name,
        iiko_product_name=iiko_product_name,
    )
    
    return generate_act(data, output_dir)
